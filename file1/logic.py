import openpyxl
from django.forms.models import model_to_dict
from datetime import datetime
from .models import Year_Program, Zadania, Recomendation, Raspisanie, Places, Teachers, Groups, Clients
from django.shortcuts import get_object_or_404


def xlxs_loader(cur_model, xlsx_file):
    fields = cur_model._meta.fields[1::]
    """try: 
        fields+=cur_model._meta.fields_many_to_many
    except:
        pass"""

    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    object_list = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        temp_data={}
        c = 0
        flag = False
        for r in row:
            try:
                fields[c]
            except:
                break
            if not(r) and fields[c].null == False:
                flag = True
            if not(r) and fields[c].null == True:
                
                if (fields[c].get_internal_type() == "CharField" or fields[c].get_internal_type() == "TextField"):
                    temp_data[fields[c].name] = ""
                elif (fields[c].get_internal_type() == "IntegerField" or fields[c].get_internal_type() == "FloatField"):
                    temp_data[fields[c].name] = 0
                elif (fields[c].get_internal_type() == "DateField"):
                    temp_data[fields[c].name] = "2000-11-11"
                elif (fields[c].get_internal_type() == "BooleanField"):
                    temp_data[fields[c].name] = False
                elif (fields[c].get_internal_type() == "ForeignKey"):
                    temp_data[fields[c].name] = None
            else:
                if (fields[c].get_internal_type() == "ForeignKey"):
                    try:
                        a = int(r)
                        get_object_or_404(fields[c].remote_field.model, pk=a)
                    except:
                        temp_data[fields[c].name] = None
                        continue
                    temp_data[fields[c].name + "_id"] = r
                elif (fields[c].get_internal_type() == "DateField"):
                    if (type(r) == str):
                        try:
                            temp_data[fields[c].name] = str(datetime.strptime(r, '%d.%m.%Y').date())
                        except:
                            temp_data[fields[c].name] = None
                    elif (type(r) == datetime):
                        temp_data[fields[c].name] = str(r.date())
                elif (fields[c].get_internal_type() == "BooleanField"):
                    temp_data[fields[c].name] = bool(r)
                else:
                    temp_data[fields[c].name] = r
            c+=1
        if (flag):
            continue
        temp_model = cur_model(**temp_data)
        object_list.append(temp_model)
  
    
    tres = cur_model.objects.bulk_create(object_list, ignore_conflicts=True)
    res = []
    for l in tres:
        res.append(model_to_dict(l))
    return res


def xlxs_loader_year_program(xlsx_file):
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    object_list = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        temp_data={}
        try:
            if (type(row[0]) == float):
                temp_data["iexact_year"] = str(int(row[0]))
            else:
                temp_data["iexact_year"] = str(row[0])
            temp_data["week_number"] = row[1]
            temp_data["idformat"] = row[2]
        except:
            continue
        if (row[3] != "" and row[3] != None):
            temp_data["raska_online"] = row[3]
        if (row[4] != "" and row[4] != None):
            temp_data["task_str_id"] = row[4]
            
            try:
                temp_data["task"] = Zadania.objects.get(task_id=row[4])
            except:
                pass

            try:
                temp_data["recom"] = Recomendation.objects.get(task_id=row[4])
            except:
                pass

        if (row[5] != "" and row[5] != None):
            temp_data["dz_str_id"] = row[5]
            try:
                temp_data["dz"] = Zadania.objects.get(task_id=row[5])
            except:
                pass
        temp_model = Year_Program(**temp_data)
        object_list.append(temp_model)
    tres = Year_Program.objects.bulk_create(object_list, ignore_conflicts=True)
    res = []
    for l in tres:
        res.append(model_to_dict(l))
    return res

def xlxs_loader_raspisanie(xlsx_file):
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    object_list = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if not(row[0]):
            continue
        temp_data={}
        temp_data["f"] = str(row[0])
        temp_data["week_day"] = str(row[1])
        temp_data["time"] = str(row[2])
        temp_data["format_plan"] = str(row[3])
        if (type(row[4]) == float):
            temp_data["program_year"] = str(int(row[4]))
        else:
            temp_data["program_year"] = str(row[4])
        if (type(row[5]) == float):
            temp_data["kabinet"] = str(int(row[5]))
        else:
            temp_data["kabinet"] = str(row[5])

        temp_data["gruop_type"] = str(row[6])
        temp_data["zoom_link"] = str(row[7])
        temp_data["id_zoom"] = str(row[8])
        temp_data["pass_zoom"] = str(row[9])
        temp_data["journal"] = str(row[10])
        temp_data["start_date"] = row[11]
        if (row[12]):
            temp_data["id"] = str(row[12])
        if (row[13]):
            temp_data["start_program"] = int(row[13])
        temp_data["fict_group"] = bool(row[14])
        temp_data["end_date"] = row[15]
        temp_place = Places.objects.get(place_id=row[16])
        temp_data["place"] = temp_place
        if (row[17]):
            temp_data["otchisl"] = int(row[17] * 100)
        temp_data["where_pay"] = str(row[18])
        if (row[19]):
            temp_data["tax"] = int(row[19] * 100)
        if (row[20]):
            temp_data["payment_per_class"] = int(row[20])

        if (row[21]):
            temp_data["payment_per_month"] = int(row[21])
        if (row[22]):
            temp_data["extra_payment"] = int(row[22] * 100)
        temp_data["mos_link"] = str(row[23])
        if (row[24]):
            temp_pedagog_table = Teachers.objects.get(teacher_id=str(row[24]))
            temp_data["pedagog_table"] = temp_pedagog_table
        if (row[25]):
            temp_pedagog_fact = Teachers.objects.get(teacher_id=str(row[25]))
            temp_data["pedagog_fact"] = temp_pedagog_fact
        if (row[26]):
            temp_data["dop_otchisl_dekart"] = int(row[26] * 100)
        

        
        

        temp_model = Raspisanie(**temp_data)
        object_list.append(temp_model)
    tres = Raspisanie.objects.bulk_create(object_list, ignore_conflicts=True)
    res = []
    for l in tres:
        res.append(model_to_dict(l))
    return res

def xlxs_loader_groups(xlsx_file):
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    object_list = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if not(row[5] and row[9]):
            continue
        temp_data={}
        temp_data["where_to"] = row[0]
        temp_data["zachislenie"] = row[1]
        temp_data["otchisl_table"] = row[2]
        temp_data["otchisl_journal"] = row[3]
        temp_data["comments"] = row[4]
        
        temp_id_group_fact = Raspisanie.objects.get(id=int(row[5][2:]))
        temp_data["id_group_fact"] = temp_id_group_fact

        temp_data["where_to_pay"] = row[6]

        temp_client_id = Clients.objects.get(special_id=str(row[7]))
        temp_data["client_id"] = temp_client_id

        temp_data["schet_num"] = str(row[8])

        temp_id_group_table = Raspisanie.objects.get(id=int(row[9][2:]))
        temp_data["id_group_table"] = temp_id_group_table

        temp_data["id"] = row[10]
        temp_model = Groups(**temp_data)
        object_list.append(temp_model)
    tres = Groups.objects.bulk_create(object_list, ignore_conflicts=True)
    res = []
    for l in tres:
        res.append(model_to_dict(l))
    return res